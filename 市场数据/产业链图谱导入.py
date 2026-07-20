#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""产业链图谱导入.py — 批量转换投喂的产业链xlsx(#040)

投喂目录: _学习/投喂/产业链/*.xlsx
  结构要求(自动探测,不合格单文件报错不中断批次):
  - 关系sheet: 三列 [节点名称|关系|节点名称], 关系∈{包含,上游,下游,中游,相关}
  - 公司sheet: 五列 [节点名称|公司简称|公司代码|产品营收|营收占比]
输出(市场数据根):
  - 产业链图谱_{板块}.json   板块=文件名去掉"产业链"后缀; 结构同#039首例
  - 产业链图谱_索引.json     各链条统计+快照日,agent一眼知道库里有哪些链
  - 产业链图谱_公司反查总表.json  跨链条合并反查: 6位代码→{简称,链条:{板块:[节点+占比]}}
幂等: 元信息.源文件md5未变则跳过(--force强制重建); 首次导入日期保留
用法: python3 产业链图谱导入.py [--force]
规矩: 图谱=证据库非模板主库; 入产业链模板.json门槛见 _agent规格/11_产业逻辑agent.md
"""
import sys, os, json, glob, hashlib, datetime, collections

BASE = os.path.dirname(os.path.abspath(__file__))
FEED = os.path.join(BASE, '_学习', '投喂', '产业链')
REL_SET = {'包含', '上游', '下游', '中游', '相关'}
TODAY = datetime.date.today().strftime('%Y-%m-%d')

def md5_of(path):
    h = hashlib.md5()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1 << 20), b''):
            h.update(chunk)
    return h.hexdigest()

def detect_sheets(wb):
    """按表头/内容探测关系sheet与公司sheet,不依赖sheet名"""
    rel_ws = co_ws = None
    for ws in wb.worksheets:
        head = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not head:
            continue
        cells = [str(c) if c else '' for c in head]
        if co_ws is None and any('公司' in c for c in cells) and any('代码' in c for c in cells):
            co_ws = ws
            continue
        if rel_ws is None:
            row2 = next(ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
            if (len(cells) >= 3 and '关系' in ''.join(cells)) or \
               (row2 and len(row2) >= 3 and str(row2[1]) in REL_SET):
                rel_ws = ws
    return rel_ws, co_ws

def convert(path, force=False):
    import openpyxl
    name = os.path.splitext(os.path.basename(path))[0]
    sector = name[:-3] if name.endswith('产业链') else name
    out_path = os.path.join(BASE, f'产业链图谱_{sector}.json')
    fp = md5_of(path)
    first_import = TODAY
    if os.path.exists(out_path):
        try:
            old_meta = json.load(open(out_path, encoding='utf-8')).get('_元信息', {})
            first_import = old_meta.get('首次导入', old_meta.get('导入日期', TODAY))
            if not force and old_meta.get('源文件md5') == fp:
                return sector, 'skip', json.load(open(out_path, encoding='utf-8'))['_元信息']['统计']
        except Exception:
            pass
    wb = openpyxl.load_workbook(path, read_only=True)
    rel_ws, co_ws = detect_sheets(wb)
    if co_ws is None:
        raise ValueError('探测不到公司sheet(需含"公司"与"代码"表头)')

    adj = {}
    if rel_ws is not None:
        for r in rel_ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0] or len(r) < 3 or not r[2]:
                continue
            a, rel, b = str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip()
            adj.setdefault(a, {}).setdefault(rel, [])
            if b not in adj[a][rel]:
                adj[a][rel].append(b)

    # (节点,代码)去重合并: 原表同一对可能两行(一行宽口径营收无占比+一行节点产品营收带占比),
    # 保留带营收占比的那条(占比=卡位纯度证据); 都无占比保留首条。丢弃占比行=历史bug,勿回退。
    merged = {}
    for r in co_ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0] or len(r) < 3 or not r[2]:
            continue
        node, nm, code = str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip()
        rev = r[3] if len(r) > 3 else None
        pct = r[4] if len(r) > 4 else None
        rev_r = round(rev, 2) if isinstance(rev, (int, float)) else None
        pct_r = round(pct, 2) if isinstance(pct, (int, float)) else None
        key = (node, code)
        if key in merged and not (pct_r is not None and merged[key][3] is None):
            continue
        merged[key] = (node, nm, code, pct_r, rev_r)
    node2co = collections.defaultdict(list)
    co_rev = {}
    for node, nm, code, pct_r, rev_r in merged.values():
        bare = code.split('.')[0]
        node2co[node].append({"代码": bare, "全代码": code, "简称": nm,
                              "产品营收": rev_r, "营收占比%": pct_r})
        e = co_rev.setdefault(bare, {"简称": nm, "全代码": code, "节点": []})
        e["节点"].append({"节点": node, "营收占比%": pct_r})
    for node in node2co:
        node2co[node].sort(key=lambda x: (-(x["营收占比%"] if x["营收占比%"] is not None else -1),
                                          -(x["产品营收"] or 0)))
    n_pct = sum(1 for v in node2co.values() for x in v if x["营收占比%"] is not None)
    stats = {"关系数": sum(len(v) for d in adj.values() for v in d.values()),
             "节点数": len(set(list(adj.keys()) + [b for d in adj.values() for v in d.values() for b in v]) |
                           set(node2co.keys())),
             "公司条目": sum(len(v) for v in node2co.values()),
             "去重公司数": len(co_rev), "带营收占比条目": n_pct}
    out = {"_元信息": {
              "来源": f"投喂xlsx: {os.path.basename(path)}",
              "首次导入": first_import, "导入日期": TODAY, "源文件md5": fp,
              "数据时效警示": "报告期与营收单位原表未标注(量级推测百万元);引用标'图谱快照',现势业绩回F10/公告核验。",
              "定位": "11号产业逻辑agent辅助证据库(非模板主库):卡位纯度复核/环节补全线索/待归位股反查。",
              "入模板门槛": "营收占比≥30%且人工核实主业卡位才提请进 产业链模板.json;泛上游通用供应商节点默认不入。",
              "统计": stats},
           "关系图谱": adj, "节点公司": dict(node2co), "公司反查": co_rev}
    tmp = out_path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=1)
    os.replace(tmp, out_path)
    return sector, 'ok', stats

def rebuild_master():
    """从全部 产业链图谱_*.json 重建索引+跨链公司反查总表"""
    index, master = {}, {}
    for p in sorted(glob.glob(os.path.join(BASE, '产业链图谱_*.json'))):
        base = os.path.basename(p)
        if base in ('产业链图谱_索引.json', '产业链图谱_公司反查总表.json'):
            continue
        sector = base[len('产业链图谱_'):-len('.json')]
        try:
            g = json.load(open(p, encoding='utf-8'))
        except Exception as e:
            print(f'  [警告] {base} 读取失败跳过: {e}')
            continue
        meta = g.get('_元信息', {})
        index[sector] = {"统计": meta.get('统计'), "导入日期": meta.get('导入日期'),
                         "首次导入": meta.get('首次导入'), "来源": meta.get('来源')}
        for code, info in g.get('公司反查', {}).items():
            m = master.setdefault(code, {"简称": info.get('简称'), "全代码": info.get('全代码'), "链条": {}})
            m["链条"][sector] = info.get('节点', [])
    idx_out = {"_说明": "产业链图谱库索引(产业链图谱导入.py自动维护,勿手改)",
               "更新日期": TODAY, "链条数": len(index), "链条": index}
    with open(os.path.join(BASE, '产业链图谱_索引.json'), 'w', encoding='utf-8') as f:
        json.dump(idx_out, f, ensure_ascii=False, indent=1)
    mst_out = {"_说明": "跨链条公司反查总表(6位代码→所在全部链条节点;涨停待归位第一查询位)",
               "更新日期": TODAY, "覆盖公司数": len(master), "反查": master}
    with open(os.path.join(BASE, '产业链图谱_公司反查总表.json'), 'w', encoding='utf-8') as f:
        json.dump(mst_out, f, ensure_ascii=False, indent=1)
    return len(index), len(master)

def main():
    force = '--force' in sys.argv
    os.makedirs(FEED, exist_ok=True)
    files = sorted(glob.glob(os.path.join(FEED, '*.xlsx')))
    files = [f for f in files if not os.path.basename(f).startswith('~$')]
    print(f'投喂目录: {FEED} | 待处理xlsx: {len(files)} | force={force}')
    ok = skip = err = 0
    for p in files:
        try:
            sector, st, stats = convert(p, force)
            if st == 'skip':
                skip += 1
                print(f'  [跳过] {sector}: 源文件未变')
            else:
                ok += 1
                print(f'  [转换] {sector}: {json.dumps(stats, ensure_ascii=False)}')
        except Exception as e:
            err += 1
            print(f'  [失败] {os.path.basename(p)}: {e}')
    n_chain, n_co = rebuild_master()
    print(f'完成: 转换{ok} 跳过{skip} 失败{err} | 索引{n_chain}条链 | 反查总表{n_co}家公司')
    if err:
        sys.exit(1)

if __name__ == '__main__':
    main()
